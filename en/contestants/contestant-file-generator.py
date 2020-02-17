#-------------------------------------------------------------
#| to rewrite HTML (F{}.html) please edited in template.html |
#-------------------------------------------------------------

#------------------------------------------------------------------
#| to update contestants (dictionary) you should do the following |
#      1. run generateData() in this file                         |
#      2. copy data from collect_data.txt and paste to this file  |
#      3. run generateFile("th") for update TH HTML files         |
#          or generateFile("en") for update EN HTML files         |
#------------------------------------------------------------------

import os;
import openpyxl as xl

n = 3

contestants = {
    'th' : {
        #contestant number 1
        1 : {
            "number" : "01",
            "concept" : "concept 1",
            "name" : "designer name 1",
        },
        #contestant number 2
        2 : {
            "number" : "02",
            "concept" : "concept 2",
            "name" : "designer name 2",
        },
        #contestant number 3
        3 : {
            "number" : "03",
            "concept" : "concept 3",
            "name" : "designer name 3",
        },
    },
    'en' : {
        #contestant number 1
        1 : {
            "number" : "05",
            "concept" : "concept 5",
            "name" : "designer name 5",
        },
        #contestant number 2
        2 : {
            "number" : "06",
            "concept" : "concept 6",
            "name" : "designer name 6",
        },
        #contestant number 3
        3 : {
            "number" : "number(en)",
            "concept" : "concept(en)",
            "name" : "name(en)",
        },
    },
}


# read Sheet Excel to create data in dictionary form
# to generate contestants data in dictionary 
# after generate ditionary you should copy collect_data.txt to this file
def generateData(lang):
    file = lang + '/contestants/contestants_sheet.xlsx'
    wb = xl.load_workbook(file)
    sh = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    language = ["'th'","'en'"]
    code = ""
    code += ("contestants = {") + "\n"
    #-------------------------
    for k in range(len(language)):
        code += (" "*4 + language[k] + " : {") + "\n"
        for i in range(1,n+1):
            code += (" "*8 + "#contestant number %d" % i) + "\n"
            code += (" "*8 + "%d : {" % i) + "\n"
            code += (" "*12 + '"number" : "{}",'.format(sh.cell(row=i+1+n*k+k,column=1).value)) + "\n"
            code += (" "*12 + '"concept" : "{}",'.format(sh.cell(row=i+1+n*k+k,column=2).value)) + "\n"
            code += (" "*12 + '"name" : "{}",'.format(sh.cell(row=i+1+n*k+k,column=3).value)) + "\n"
            code += (" "*8 + "},") + "\n"
        code += (" "*4 + "},") + "\n"
    #--------------------------
    code += ("}") + "\n"
    print(code)
    # contain data in collect_data.txt
    f = open(lang + "/contestants/collect_data.txt","w")
    f.write(code)
    f.close()

# to delete all contestants .html file
def deleteAllHtmlFile():
    for i in range(1,n+1):
        fname = "contestants/F{}.html".format(i)
        if os.path.exists(fname):
            os.remove(fname)
            print(fname,"was deleted")
        else:
            print(fname,"does not exist")

# to delete specific content in file 
def deleteContent(fName):
    with open(fName, "w"):
        pass

# to overwrite specific .html file with contestants data in ditionary 
def genreateHtml(html,index,lang):
    DATA = contestants[lang][index]
    html = html.format(index, #{0}
    index, #{1}
    index, #{2}
    index, #{3}
    DATA["concept"], #{4}
    DATA["name"]) #{5}
    return html

# to generate All .html file
# for ver.th please call generateFile("th")
# for ver.en please call generateFile("en")
def generateFile(lang):
    contents = ""
    t = open(lang + "/contestants/template.html","r");
    if t.mode == "r":
            contents = t.read()

    for i in range(1,n+1):
        fname = lang + "/contestants/S{}.html".format(i)
        f = open(fname,"w")
        f.write(genreateHtml(contents,i,lang))
        f.close()

if __name__ == "__main__":
    #deleteAllHtmlFile()
    print("PLEASE SELECT LANGUAGE (th/en) : ",end="")
    s = input()
    if s=="th":
        generateData("th")
        generateFile("th")
    else :
        generateData("en")
        generateFile("en")
    print("done")